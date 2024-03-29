import shutil
import openpyxl as xl
import os
import glob
import re
from copy import copy
from win32com.client import Dispatch

path = os.getcwd()
xml_files = glob.glob(os.path.join(path, "*.xml"))
excel_name = "Testing"
exclusion_list = ("Blank", "Kalib.blank", "Std1", "Std2", "Std3", "QC", "qc", "Vesi", "vesi") # List of labels to filter out

process_samples = ("AgPbR_", "ZnR_", "SR_", "M_", "J_", "ZnJ_", "AgPbJ_")
courier_samples = (" AgPbR", " ZnR", " SR", " M", " J", " ZnJ", " AgPbJ")
qc = ["Prep", "Pulp", "GBM", "Reag.blank", "GMR", "OREAS", "GSB", "Pyrref", "Pyr-ref", "SR-Ref"]

not_geological = []
value_col = 3 # column number of the column with the names

def convert_xls_to_xlsx(oldName:str, newName:str):
    oldName = os.path.abspath(oldName)
    newName = os.path.abspath(newName)
    xlApp = Dispatch("Excel.Application")
    wb = xlApp.Workbooks.Open(oldName)
    wb.SaveAs(newName,51)
    wb.Close(True)   

def move_items(from_sheet, to_sheet, list):
    # Move data from a sheet to another if column contains string in list
    sheet1_row = 2
    sheet2_row = 0          
    for row in from_sheet.iter_rows(min_row = 3, min_col = value_col, max_col = value_col): 
        sheet1_row +=1
        for cell in row:
            if any(value in cell.value for value in list) and "Pulp".lower() not in cell.value.lower() and "Prep".lower() not in cell.value.lower():
                if list == courier_samples and "sr-ref" in cell.value.lower():  # fix for SR-ref samples being placed in courier sample list                 
                    break
                not_geological.append(cell.value)
                for row2 in from_sheet.iter_rows(min_row = sheet1_row, max_row = sheet1_row, min_col = 1):
                    sheet2_row +=1
                    for cell in row2:
                        to_sheet.cell(row=sheet2_row, column = cell.col_idx).value = cell.value
                        if cell.has_style:
                                to_sheet.cell(row=sheet2_row, column = cell.col_idx)._style = copy(cell._style)
def move_sorters(from_sheet, to_sheet):
    sheet1_row = 2
    sheet2_row = 0
    for row in from_sheet.iter_rows(min_row = 3, min_col = value_col, max_col = value_col):
        sheet1_row +=1
        for cell in row:
            if "_SO_" in cell.value:
                not_geological.append(cell.value)
                for row2 in from_sheet.iter_rows(min_row = sheet1_row, max_row = sheet1_row, min_col = 1):
                    sheet2_row +=1
                    for cell in row2:
                        to_sheet.cell(row=sheet2_row, column = cell.col_idx).value = cell.value
                        if cell.has_style:
                                to_sheet.cell(row=sheet2_row, column = cell.col_idx)._style = copy(cell._style)
def move_qc(from_sheet, to_sheet):
    sheet1_row = 2
    sheet2_row = 0
    lower_qc = [value.lower() for value in qc] # convert list to lower case to compare with lower case cell values
    indices = []
    for row in from_sheet.iter_rows(min_row = 3, min_col = value_col, max_col = value_col):
        
        sheet1_row +=1
        for cell in row:
            if any(value in cell.value.lower() for value in lower_qc):
                not_geological.append(cell.value)
                # if prep or pulp in cell, loops through the values again and moves the original value before the duplicate
                if "prep" in cell.value.lower() or "pulp" in cell.value.lower():
                    non_dup = cell.value.lower().split("_", 1)
                    if "pulp" in cell.value.lower():
                        non_dup_split = non_dup[1].split(" pulp")
                        non_dup_final = "_" + non_dup_split[0]                                                                           
                    else:                        
                        non_dup_split = non_dup[1].split(" prep")
                        non_dup_final = "_" + non_dup_split[0]
                        
                    this_row = 2    
                    break_loop = False
                    for row in from_sheet.iter_rows(min_row = 3, min_col = value_col, max_col = value_col):
                        this_row += 1
                        for cell in row:
                            if non_dup_final in cell.value.lower() and "prep" not in cell.value.lower() and "pulp" not in cell.value.lower():                             
                                if cell.row not in indices:                                    
                                    for row in from_sheet.iter_rows(min_row = this_row, max_row = this_row, min_col = 1):
                                        sheet2_row += 1
                                        for cell in row:
                                            to_sheet.cell(row=sheet2_row, column = cell.col_idx).value = cell.value
                                            if cell.has_style:
                                                    to_sheet.cell(row=sheet2_row, column = cell.col_idx)._style = copy(cell._style)
                                    indices.append(cell.row) # add index value to list so that same value is not added again
                                    break_loop = True
                                break
                            if break_loop:
                                break
                        if break_loop:
                            break                                     
                
                for row2 in from_sheet.iter_rows(min_row = sheet1_row, max_row = sheet1_row, min_col = 1): # adds the pulp/prep value afterwards
                    sheet2_row +=1
                    for cell in row2:
                        to_sheet.cell(row=sheet2_row, column = cell.col_idx).value = cell.value
                        if cell.has_style:
                                to_sheet.cell(row=sheet2_row, column = cell.col_idx)._style = copy(cell._style)

num = 0
for f in xml_files:
    num += 1
    name_stripped = f.strip(".xml")
    excel_file = f"{name_stripped}.xlsx"
    convert_xls_to_xlsx(f, excel_file)
    wb = xl.load_workbook(excel_file)
    ws1 = wb.active
    wb.create_sheet("Sorted")
    ws2 = wb["Sorted"]

    for row in ws1.iter_rows(min_row = 1, max_row = 2, min_col = 1, max_col = 14):
        for cell in row:
            ws2[f"{cell.coordinate}"] = cell.value
            if cell.has_style:
                    ws2[f"{cell.coordinate}"]._style = copy(cell._style)

    for row in ws1.iter_rows(min_row = 3, min_col = 1, max_col = 14):   
        for cell in row:
            ws2[f"{cell.coordinate}"] = cell.value
            if cell.has_style:
                    ws2[f"{cell.coordinate}"]._style = copy(cell._style)

    # Remove rows with a cell that has a string from exclusion list
    for i in reversed(range(2, ws2.max_row+1)):
        if ws2.cell(row= i, column = value_col).value in exclusion_list:
            ws2.delete_rows(i)
    
    # Strip g and ml units from a few columns

    for row in ws2.iter_rows(min_row = 3, min_col= 6, max_col = 6):
        for cell in row:
            txt = cell.value
            try:
                stripped = txt.strip(" g")
            except(AttributeError):
                continue
            try:
                cell.value = float(stripped)
            except(ValueError):
                continue

    for row in ws2.iter_rows(min_row = 3, min_col= 7, max_col = 7):
        for cell in row:
            txt = cell.value
            try:
                stripped = txt.strip(" ml")
            except(AttributeError):
                continue
            try:
                cell.value = float(stripped)
            except(ValueError):
                continue

    wb.create_sheet("Process")
    wb.create_sheet("Courier")  
    wb.create_sheet("Sorter")
    wb.create_sheet("QC")  
    wb.create_sheet("Final Sorted")

    ws3 = wb["Process"]
    ws4 = wb["Courier"]
    ws5 = wb["Sorter"]
    ws6 = wb["QC"]
    ws7 = wb["Final Sorted"]                            

    move_items(ws2,ws3, process_samples)
    move_items(ws2,ws4, courier_samples)
    move_sorters(ws2, ws5)
    move_qc(ws2, ws6)

    for row in ws1.iter_rows(min_row = 1, max_row = 2, min_col = 1, max_col = 13):
        for cell in row:
            ws7[f"{cell.coordinate}"] = cell.value

    row_num = 2
    row_idx = 2
    # Add geologicals to final sheet
    for row in ws2.iter_rows(min_row = 3, min_col = value_col, max_col = value_col):
        row_num += 1
        for cell in row:
            if cell.value not in not_geological:               
                for row in ws2.iter_rows(min_row = row_num, max_row = row_num, min_col = 1, max_col = 14):
                    row_idx += 1
                    for cell in row:
                        ws7.cell(row = row_idx, column = cell.col_idx).value = cell.value
                        if cell.has_style:
                            ws7.cell(row = row_idx, column = cell.col_idx)._style = copy(cell._style)
    row_idx += 2

    def add_non_geos(from_sheet, to_sheet):
        global row_idx
        is_empty = False
        if from_sheet["A1"].value == None:
            is_empty = True
        if is_empty == False: # loop through rows only if sheet was not empty
            indicator = from_sheet.title + ":"
            to_sheet.cell(row = row_idx, column = 1).value = indicator
            for row in from_sheet.iter_rows(min_row = 1, min_col = 1, max_col = 14):
                row_idx += 1 
                for cell in row:
                    to_sheet.cell(row = row_idx, column = cell.col_idx).value = cell.value            
                    if cell.has_style:
                        to_sheet.cell(row = row_idx, column = cell.col_idx)._style = copy(cell._style) 
        if is_empty == False:            
            row_idx += 2

    add_non_geos(ws3, ws7)
    add_non_geos(ws4, ws7)
    add_non_geos(ws5, ws7)
    add_non_geos(ws6, ws7)
    
    wb.remove(ws3)
    wb.remove(ws4)
    wb.remove(ws5)
    wb.remove(ws6)

    wb.save(excel_file)
    wb.close()

def move_file(xml_file):
    dir_name = "xmls"
    xmls_dir = os.path.join(path, dir_name)
    xml = os.path.join(path, xml_file)
    if not os.path.exists(xmls_dir):      
        os.mkdir(xmls_dir)
    shutil.move(xml, xmls_dir)

for f in xml_files:        
    move_file(f)